import requests
from bs4 import BeautifulSoup
import openpyxl

# Set up input and output file paths
input_file = 'input.xlsx'
output_file = 'Douglas_output.xlsx'

# Load input file and output file
input_wb = openpyxl.load_workbook(input_file)
output_wb = openpyxl.Workbook()

# Get the active worksheet from the input file
input_ws = input_wb.active

# Create a new worksheet in the output file
output_ws = output_wb.active
output_ws.title = 'Results'

# Write the headers to the output file
output_ws['A1'] = 'EAN'
output_ws['B1'] = 'Brand'
output_ws['C1'] = 'Product Name'
output_ws['D1'] = 'Size'
output_ws['E1'] = 'Price'

for row in input_ws.iter_rows(min_row=2, values_only=True):
    ean = str(row[0]).strip()
    url = f"https://www.douglas.nl/nl/search?q={ean}"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    # find product brand and name
    product_header = soup.find(class_='product-detail-header__group')
    if product_header:     
        brand_element = product_header.find(class_='brand-logo__text brand-logo__text--dynamic')
        name_element = product_header.find(class_='second-line')
        if brand_element and name_element:
            brand = brand_element.text.strip()
            name = name_element.text.strip()

    # find product size and price
    product_tiles = soup.find_all(class_='product-detail__variant-row product-detail__variant-row--spread-content')
    if product_tiles:
        for product_tile in product_tiles:
            size_element = product_tile.find(class_='product-detail__variant-name')
            price_element = product_tile.find(class_='product-price__price')
            if  size_element and price_element:
                size = size_element.text.strip()
                price_str = price_element.text.strip().replace('€', '').replace('\xa0', '').replace(',', '.')
                price = float(price_str)
                output_ws.append([ean, brand, name, size, price])
            else:
                output_ws.append([ean, "No name or price found", ""])
    else:
        output_ws.append([ean, "No products found", ""])

# Save the output file
output_wb.save(output_file)