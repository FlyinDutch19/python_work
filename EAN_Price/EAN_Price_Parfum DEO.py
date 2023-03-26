import requests
from bs4 import BeautifulSoup
import openpyxl

# Set up input and output file paths
input_file = 'input.xlsx'
output_file = 'Parfum DEO_output.xlsx'

# Load input file and output file
input_wb = openpyxl.load_workbook(input_file)
output_wb = openpyxl.Workbook()

# Get the active worksheet from the input file
input_ws = input_wb.active

# Create a new worksheet in the output file
output_ws = output_wb.active
output_ws.title = 'Results'

# Write the headers to the output file
output_ws['A1'] = 'EAN'
output_ws['B1'] = 'Product Name'
output_ws['C1'] = 'Price'

# Loop over each row in the input file
for row in input_ws.iter_rows(min_row=2, values_only=True):
    ean = str(row[0]).strip()
    url = f"https://www.parfumdeo.nl/catalogsearch/result/?q={ean}"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    product_tile = soup.find(class_='column main')
    if product_tile:
            name_element = product_tile.find(class_='page-title')
            price_element = product_tile.find(class_='price')
            if name_element and price_element:
                name = name_element.text.strip()
                price_str = price_element.text.strip().replace(',', '.').replace('€', '')
                if price_str:
                    price = float(price_str)
                else:
                    price = None
                output_ws.append([ean, name, price])
            else:
                output_ws.append([ean, "No name or price found", None])
    else:
        output_ws.append([ean, "No products found", None])

# Save the output file
output_wb.save(output_file)
