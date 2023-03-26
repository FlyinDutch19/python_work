import requests
from bs4 import BeautifulSoup
import openpyxl

# Set up input and output file paths
input_file = 'input.xlsx'
output_file = 'Bol&DEO_output.xlsx'

# Load input file and create output file
input_wb = openpyxl.load_workbook(input_file)
output_wb = openpyxl.Workbook()

# Create a new worksheet in the output file
output_ws = output_wb.active
output_ws.title = 'Results'

# Write the headers to the output file
output_ws['A1'] = 'EAN'
output_ws['B1'] = 'Bol.com Name'
output_ws['C1'] = 'Bol.com Price'
output_ws['D1'] = 'Parfum DEO Name'
output_ws['E1'] = 'Parfum DEO Price'

# Define a function to scrape data from Bol.com
def scrape_bol(ean):
    url = f"https://www.bol.com/nl/nl/s/?searchtext={ean}"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    product_tile = soup.find(class_='product-item__content')
    if product_tile:
        name_element = product_tile.find(class_='product-title px_list_page_product_click list_page_product_tracking_target')
        price_element = product_tile.find(class_='promo-price')
        if name_element and price_element:
            name = name_element.text.strip()
            if price_element is not None:
                price_text = price_element.text.strip().replace(' ', ',').replace(',,', ',')
                # Remove any non-numeric characters from the price string
                price_text = ''.join(c for c in price_text if c.isdigit() or c == ',')
                price = float(price_text.replace(',', '.'))
            else:
                price = None
            return name, price
        else:
            return "No name or price found", None
    else:
        return "No products found", None

# Define a function to scrape data from Parfum DEO
def scrape_parfumdeo(ean):
    url = f"https://www.parfumdeo.nl/catalogsearch/result/?q={ean}"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    product_tile = soup.find(class_='column main')
    if product_tile:
        name_element = product_tile.find(class_='page-title')
        price_element = product_tile.find(class_='price')
        if name_element and price_element:
            name = name_element.text.strip()
            price_str = price_element.text.strip().replace(',', '.').replace('€', '')
            if price_str:
                price = float(price_str)
            else:
                price = None
            return name, price
        else:
            return "No name or price found", None
    else:
        return "No products found", None

# Loop over each row in the input file
for row in input_wb.active.iter_rows(min_row=2, values_only=True):
    ean = str(row[0]).strip()
    bol_name, bol_price = scrape_bol(ean)
    parfumdeo_name, parfumdeo_price = scrape_parfumdeo(ean)

    # Write the data to the output file
    output_ws.append([ean, bol_name, bol_price, parfumdeo_name, parfumdeo_price])

# Save the output file
output_wb.save(output_file)
