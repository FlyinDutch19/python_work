import requests
from bs4 import BeautifulSoup
import openpyxl

# Set up input and output file paths
input_file = 'input.xlsx'
output_file = 'Bol_output.xlsx'

# Load input file and output file
input_wb = openpyxl.load_workbook(input_file)
output_wb = openpyxl.Workbook()

# Get the active worksheet from the input file
input_ws = input_wb.active

# Create a new worksheet in the output file
output_ws = output_wb.active
output_ws.title = 'Results'

# Write the headers to the output file
output_ws['A1'] = 'EAN'
output_ws['B1'] = 'Product Name'
output_ws['C1'] = 'Price'

# Loop over each row in the input file
for row in input_ws.iter_rows(min_row=2, values_only=True):
    ean = str(row[0]).strip()
    url = f"https://www.bol.com/nl/nl/s/?searchtext={ean}"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    product_tile = soup.find(class_='product-item__content')
    if product_tile:
            name_element = product_tile.find(class_='product-title px_list_page_product_click list_page_product_tracking_target')
            price_element = product_tile.find(class_='promo-price')
            if name_element and price_element:
                name = name_element.text.strip()
                if price_element is not None:
                    price = price_element.text.strip().replace(' ', ',')
                else:
                    price = ""
                output_ws.append([ean, name, price])
            else:
                output_ws.append([ean, "No name or price found", ""])
    else:
        output_ws.append([ean, "No products found", ""])


# Save the output file
output_wb.save(output_file)