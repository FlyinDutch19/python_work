import requests
import json
import openpyxl

def get_product_name(ean):
    url = f'https://api.upcitemdb.com/prod/trial/lookup?upc={ean}'
    response = requests.get(url)
    data = json.loads(response.text)
    if 'items' in data and data['items']:
        product_name = data['items'][0]['title']
        return product_name
    else:
        return "Product not found"

# Load input data from Excel file
workbook = openpyxl.load_workbook('input.xlsx')
worksheet = workbook.active
rows = worksheet.iter_rows(min_row=2, values_only=True)

# Process each row of input data
output_rows = []
for row in rows:
    ean = row[0]
    product_name = get_product_name(str(ean))
    output_rows.append((ean, product_name))

# Save output data to Excel file
output_workbook = openpyxl.Workbook()
output_worksheet = output_workbook.active
output_worksheet.append(('EAN', 'Product Name'))
for row in output_rows:
    output_worksheet.append(row)
output_workbook.save('output.xlsx')
