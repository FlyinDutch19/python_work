import openpyxl
import openai
import os
from dotenv import load_dotenv

load_dotenv()

# Load the Excel file containing English product names
wb = openpyxl.load_workbook('input.xlsx')
ws = wb.active

# Create a new Excel file for the translated product names
translated_wb = openpyxl.Workbook()
translated_ws = translated_wb.active
translated_ws.title = "Translated"

# Initialize OpenAI API key
openai.api_key = "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"

# Iterate through each row in the input Excel file and translate the product name
for row in ws.iter_rows(min_row=2):
    product_name = row[0].value
    response = openai.Completion.create(
        model="text-davinci-002",
        prompt=f"Translate this English product name to Chinese: {product_name}",
        temperature=0.7,
        max_tokens=60,
        n=1,
        stop=None,
        )
    translated_product_name = response.choices[0].text.strip()
    translated_ws.append([translated_product_name])

# Save the translated Excel file
translated_wb.save('translated.xlsx')